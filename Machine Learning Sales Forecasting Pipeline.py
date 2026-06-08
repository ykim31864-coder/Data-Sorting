from __future__ import annotations

import io
import json
import math
import re
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import numpy as np
import pandas as pd


ROOT = Path(r"C:\Users\user\Documents\Codex\2026-06-05\files-mentioned-by-the-user-ml")
OUTPUT_DIR = ROOT / "outputs"

SOURCE_FILES = {
    "data_zip": Path(r"C:\Users\user\Downloads\data (2).zip"),
    "team": Path(r"C:\Users\user\Downloads\팀별 실적 현황.xlsx"),
    "snu_items": Path(r"C:\Users\user\Downloads\서울대 품목별 실적.xlsx"),
    "account_rank": Path(r"C:\Users\user\Downloads\거래처 실적 순위.xlsx"),
    "metoject_accounts": Path(r"C:\Users\user\Downloads\메토젝트 종병 거래처 실적 순위.xlsx"),
}

FORECAST_START = pd.Timestamp("2026-05-01")
FORECAST_END = pd.Timestamp("2026-12-01")
FORECAST_MONTHS = pd.date_range(FORECAST_START, FORECAST_END, freq="MS")


def clean_number(value: object) -> float:
    if value is None or pd.isna(value):
        return np.nan
    if isinstance(value, (int, float, np.integer, np.floating)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if text in {"", "-", "nan", "None"}:
        return np.nan
    if text.endswith("%"):
        text = text[:-1]
    return float(text)


def read_xlsx_from_zip(zip_path: Path, member: str) -> pd.DataFrame:
    with zipfile.ZipFile(zip_path) as zf:
        with zf.open(member) as fh:
            return pd.read_excel(io.BytesIO(fh.read()))


def load_sources() -> dict[str, pd.DataFrame]:
    return {
        "item_yearly": read_xlsx_from_zip(SOURCE_FILES["data_zip"], "품목별 실적.xlsx"),
        "hospital_yearly": read_xlsx_from_zip(SOURCE_FILES["data_zip"], "병원별 실적.xlsx"),
        "team": pd.read_excel(SOURCE_FILES["team"]),
        "snu_items": pd.read_excel(SOURCE_FILES["snu_items"]),
        "account_rank": pd.read_excel(SOURCE_FILES["account_rank"]),
        "metoject_accounts": pd.read_excel(SOURCE_FILES["metoject_accounts"]),
    }


def parse_year(value: object) -> int | None:
    match = re.search(r"(20\d{2})", str(value))
    return int(match.group(1)) if match else None


def long_from_year_month_table(
    df: pd.DataFrame,
    source: str,
    id_cols: list[str],
    value_suffix: str = "월 실적",
) -> pd.DataFrame:
    rows = []
    month_cols = []
    for col in df.columns:
        match = re.fullmatch(r"(\d{1,2})" + re.escape(value_suffix), str(col))
        if match:
            month_cols.append((col, int(match.group(1))))

    for _, row in df.iterrows():
        year = parse_year(row.get("연도"))
        if year is None:
            continue
        ids = {col: row.get(col) for col in id_cols if col in df.columns}
        for col, month_num in month_cols:
            value = clean_number(row.get(col))
            if pd.isna(value):
                continue
            rows.append(
                {
                    "source": source,
                    **ids,
                    "month": pd.Timestamp(year=year, month=month_num, day=1),
                    "sales": value,
                }
            )
    return pd.DataFrame(rows)


def long_from_ym_columns(
    df: pd.DataFrame,
    source: str,
    id_cols: list[str],
    month_pattern: str = r"20\d{2}-\d{2}",
) -> pd.DataFrame:
    rows = []
    month_cols = [c for c in df.columns if re.fullmatch(month_pattern, str(c))]
    for _, row in df.iterrows():
        ids = {col: row.get(col) for col in id_cols if col in df.columns}
        for col in month_cols:
            value = clean_number(row.get(col))
            if pd.isna(value):
                continue
            rows.append(
                {
                    "source": source,
                    **ids,
                    "month": pd.Timestamp(str(col) + "-01"),
                    "sales": value,
                }
            )
    return pd.DataFrame(rows)


def team_long(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    actual_rows = []
    target_rows = []
    for _, row in df.iterrows():
        ids = {"본부": row.get("본부"), "부서": row.get("부서")}
        for month in range(1, 13):
            month_start = pd.Timestamp(year=2026, month=month, day=1)
            target = clean_number(row.get(f"2026-{month:02d} 목표"))
            actual = clean_number(row.get(f"2026-{month:02d} 실적"))
            if not pd.isna(target):
                target_rows.append({"source": "팀별 목표", **ids, "month": month_start, "target": target})
            if not pd.isna(actual):
                actual_rows.append({"source": "팀별 실적", **ids, "month": month_start, "sales": actual})
    return pd.DataFrame(actual_rows), pd.DataFrame(target_rows)


def aggregate_monthly(df: pd.DataFrame) -> pd.DataFrame:
    monthly = (
        df.groupby("month", as_index=False)["sales"]
        .sum()
        .sort_values("month")
        .reset_index(drop=True)
    )
    monthly["sales"] = pd.to_numeric(monthly["sales"], errors="coerce")
    return monthly


def add_features(monthly: pd.DataFrame) -> pd.DataFrame:
    out = monthly.sort_values("month").reset_index(drop=True).copy()
    out["year"] = out["month"].dt.year
    out["month_num"] = out["month"].dt.month
    out["period_idx"] = np.arange(len(out), dtype=float)
    out["sin_month"] = np.sin(2 * np.pi * out["month_num"] / 12)
    out["cos_month"] = np.cos(2 * np.pi * out["month_num"] / 12)
    for lag in [1, 2, 3, 6, 12]:
        out[f"lag_{lag}"] = out["sales"].shift(lag)
    out["rolling_3"] = out["sales"].shift(1).rolling(3).mean()
    out["rolling_6"] = out["sales"].shift(1).rolling(6).mean()
    out["yoy"] = out["sales"] / out["lag_12"] - 1
    out = out.replace([np.inf, -np.inf], np.nan)
    return out


def candidate_feature_sets(monthly: pd.DataFrame) -> list[list[str]]:
    n = len(monthly.dropna(subset=["sales"]))
    base = ["period_idx", "month_num", "sin_month", "cos_month"]
    sets = [
        base + ["lag_1", "lag_2", "lag_3", "rolling_3"],
        base + ["lag_1", "lag_2", "lag_3", "lag_6", "rolling_3", "rolling_6"],
    ]
    if n >= 18:
        sets.append(base + ["lag_1", "lag_2", "lag_3", "lag_6", "lag_12", "rolling_3", "rolling_6", "yoy"])
    return sets


@dataclass
class ModelResult:
    name: str
    model: object
    feature_cols: list[str]
    mae: float
    mape: float
    r2: float
    train_rows: int
    test_rows: int


class RidgeRegressor:
    def __init__(self, alpha: float = 1.0, log_target: bool = False, recent_weight: bool = False):
        self.alpha = alpha
        self.log_target = log_target
        self.recent_weight = recent_weight
        self.mean_: np.ndarray | None = None
        self.std_: np.ndarray | None = None
        self.coef_: np.ndarray | None = None
        self.target_cap_: float | None = None

    def fit(self, x: pd.DataFrame, y: pd.Series) -> "RidgeRegressor":
        arr = x.to_numpy(dtype=float)
        target = y.to_numpy(dtype=float)
        finite_target = target[np.isfinite(target)]
        self.target_cap_ = float(max(np.nanmax(finite_target) * 3, 1.0)) if len(finite_target) else 1.0
        if self.log_target:
            target = np.log1p(np.maximum(target, 0))
        self.mean_ = arr.mean(axis=0)
        self.std_ = arr.std(axis=0)
        self.std_[self.std_ == 0] = 1
        z = (arr - self.mean_) / self.std_
        design = np.column_stack([np.ones(len(z)), z])
        weights = np.ones(len(z))
        if self.recent_weight:
            weights = np.linspace(0.55, 1.45, len(z))
        weighted_design = design * np.sqrt(weights)[:, None]
        weighted_target = target * np.sqrt(weights)
        penalty = np.eye(design.shape[1]) * self.alpha
        penalty[0, 0] = 0
        self.coef_ = np.linalg.pinv(weighted_design.T @ weighted_design + penalty) @ weighted_design.T @ weighted_target
        return self

    def predict(self, x: pd.DataFrame) -> np.ndarray:
        if self.mean_ is None or self.std_ is None or self.coef_ is None:
            raise ValueError("Model has not been fitted.")
        arr = x.to_numpy(dtype=float)
        z = (arr - self.mean_) / self.std_
        design = np.column_stack([np.ones(len(z)), z])
        pred = design @ self.coef_
        if self.log_target:
            max_log = np.log1p(self.target_cap_ if self.target_cap_ is not None else 1.0)
            pred = np.clip(pred, 0, max_log)
            pred = np.expm1(pred)
        return np.maximum(pred, 0.0)


def mae(y_true: pd.Series, y_pred: np.ndarray) -> float:
    return float(np.mean(np.abs(y_true.to_numpy(dtype=float) - y_pred)))


def mape(y_true: pd.Series, y_pred: np.ndarray) -> float:
    actual = y_true.to_numpy(dtype=float)
    mask = actual != 0
    if not mask.any():
        return float("nan")
    return float(np.mean(np.abs((actual[mask] - y_pred[mask]) / actual[mask])))


def r2_score(y_true: pd.Series, y_pred: np.ndarray) -> float:
    actual = y_true.to_numpy(dtype=float)
    ss_res = np.sum((actual - y_pred) ** 2)
    ss_tot = np.sum((actual - actual.mean()) ** 2)
    return float(1 - ss_res / ss_tot) if ss_tot else float("nan")


def evaluate_ridge(monthly: pd.DataFrame) -> ModelResult:
    featured = add_features(monthly)
    candidates = []
    for feature_cols in candidate_feature_sets(featured):
        model_df = featured.dropna(subset=feature_cols + ["sales"]).copy()
        if len(model_df) < 6:
            continue
        test_size = min(4, max(2, len(model_df) // 4))
        train = model_df.iloc[:-test_size]
        test = model_df.iloc[-test_size:]
        for alpha in [0.5, 1.0, 3.0]:
            for log_target in [False, True]:
                for recent_weight in [False, True]:
                    name = f"{'Weighted' if recent_weight else ''}{'Log' if log_target else ''}Ridge(alpha={alpha:g})"
                    model = RidgeRegressor(alpha=alpha, log_target=log_target, recent_weight=recent_weight)
                    model.fit(train[feature_cols], train["sales"])
                    pred = model.predict(test[feature_cols])
                    if not np.isfinite(pred).all():
                        continue
                    result = ModelResult(
                        name=name,
                        model=model,
                        feature_cols=feature_cols,
                        mae=mae(test["sales"], pred),
                        mape=mape(test["sales"], pred),
                        r2=r2_score(test["sales"], pred) if len(test) > 1 else np.nan,
                        train_rows=len(train),
                        test_rows=len(test),
                    )
                    candidates.append((result.mape, result.mae, result))
    if not candidates:
        raise ValueError("Not enough monthly observations for Ridge forecasting.")
    best = sorted(candidates, key=lambda x: (math.inf if pd.isna(x[0]) else x[0], x[1]))[0][2]
    model_df = featured.dropna(subset=best.feature_cols + ["sales"]).copy()
    best.model.fit(model_df[best.feature_cols], model_df["sales"])
    best.train_rows = len(model_df)
    return best


def simple_md_table(df: pd.DataFrame) -> str:
    if df.empty:
        return "_데이터 없음_"
    formatted = df.copy()
    for col in formatted.columns:
        if pd.api.types.is_numeric_dtype(formatted[col]):
            formatted[col] = formatted[col].map(lambda v: "" if pd.isna(v) else f"{v:,.4f}" if abs(v) < 10 else f"{v:,.2f}")
        else:
            formatted[col] = formatted[col].astype(str)
    headers = [str(c) for c in formatted.columns]
    lines = [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join(["---"] * len(headers)) + " |",
    ]
    for row in formatted.values.tolist():
        lines.append("| " + " | ".join(str(v) for v in row) + " |")
    return "\n".join(lines)


def forecast_series(monthly: pd.DataFrame, forecast_months: Iterable[pd.Timestamp]) -> tuple[pd.DataFrame, ModelResult]:
    history = monthly[monthly["month"] < min(forecast_months)].copy()
    best = evaluate_ridge(history)
    forecasts = []
    for month in forecast_months:
        tmp = pd.concat(
            [history, pd.DataFrame([{"month": month, "sales": np.nan}])],
            ignore_index=True,
        )
        tmp = add_features(tmp)
        idx = tmp.index[-1]
        pred = float(best.model.predict(tmp.loc[[idx], best.feature_cols])[0])
        recent_values = history["sales"].tail(6)
        same_month_last_year = history.loc[history["month"].eq(month - pd.DateOffset(years=1)), "sales"]
        current_year_actuals = history.loc[
            (history["month"].dt.year == month.year) & (history["month"] < month),
            "sales",
        ]
        baseline_parts = [
            history["sales"].tail(3).mean(),
            recent_values.mean(),
            same_month_last_year.iloc[-1] if not same_month_last_year.empty else np.nan,
            current_year_actuals.mean() if not current_year_actuals.empty else np.nan,
        ]
        baseline_parts = [float(v) for v in baseline_parts if not pd.isna(v) and np.isfinite(v) and v > 0]
        if baseline_parts:
            baseline = float(np.median(baseline_parts))
            if not np.isfinite(pred):
                pred = baseline
            pred = min(max(pred, baseline * 0.35), baseline * 2.50)
        elif not np.isfinite(pred):
            pred = 0.0
        history = pd.concat(
            [history, pd.DataFrame([{"month": month, "sales": pred}])],
            ignore_index=True,
        )
        forecasts.append({"month": month, "forecast_sales": pred})
    return pd.DataFrame(forecasts), best


def forecast_groups(
    long_df: pd.DataFrame,
    group_cols: list[str],
    min_months: int,
    top_n: int,
    source_label: str,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    recent_cutoff = FORECAST_START
    group_sales = (
        long_df[long_df["month"] < recent_cutoff]
        .groupby(group_cols, dropna=False)["sales"]
        .sum()
        .sort_values(ascending=False)
        .head(top_n)
        .reset_index()
    )
    rows = []
    score_rows = []
    for _, group in group_sales.iterrows():
        mask = pd.Series(True, index=long_df.index)
        label_parts = []
        for col in group_cols:
            value = group[col]
            if pd.isna(value):
                mask &= long_df[col].isna()
                label_parts.append("")
            else:
                mask &= long_df[col].eq(value)
                label_parts.append(str(value))
        monthly = aggregate_monthly(long_df[mask])
        if len(monthly[monthly["month"] < FORECAST_START]) < min_months or monthly["sales"].sum() <= 0:
            continue
        try:
            fc, model = forecast_series(monthly, FORECAST_MONTHS)
        except Exception:
            continue
        group_name = " / ".join([p for p in label_parts if p]) or "전체"
        recent_actual = monthly[(monthly["month"] >= "2026-01-01") & (monthly["month"] <= "2026-04-01")]["sales"].sum()
        forecast_sum = fc["forecast_sales"].sum()
        for _, r in fc.iterrows():
            rows.append({"source": source_label, "group": group_name, "month": r["month"], "forecast_sales": r["forecast_sales"]})
        score_rows.append(
            {
                "source": source_label,
                "group": group_name,
                "recent_2026_1_4_actual": recent_actual,
                "forecast_2026_5_12": forecast_sum,
                "model": model.name,
                "mape": model.mape,
                "mae": model.mae,
                "r2": model.r2,
                "train_rows": model.train_rows,
                "features": ", ".join(model.feature_cols),
            }
        )
    return pd.DataFrame(rows), pd.DataFrame(score_rows)


def build_all_long(data: dict[str, pd.DataFrame]) -> dict[str, pd.DataFrame]:
    item_long = long_from_year_month_table(
        data["item_yearly"],
        "품목별 실적",
        ["병원구분", "기준규격", "브랜드", "DC브랜드"],
    )
    hospital_long = long_from_year_month_table(
        data["hospital_yearly"],
        "병원별 실적",
        ["병원구분", "병원", "브랜드", "DC브랜드", "기준규격"],
    )
    team_actual, team_target = team_long(data["team"])
    snu_long = long_from_ym_columns(
        data["snu_items"],
        "서울대 품목별 실적",
        ["병원", "기준규격", "브랜드", "DC브랜드", "구분"],
    )
    account_long = long_from_ym_columns(
        data["account_rank"],
        "거래처 실적 순위",
        ["본부", "병원", "팀", "담당자"],
    )
    meto_long = long_from_ym_columns(
        data["metoject_accounts"],
        "메토젝트 종병 거래처 실적 순위",
        ["브랜드", "본부", "병원", "팀", "담당자"],
    )
    return {
        "item_long": item_long,
        "hospital_long": hospital_long,
        "team_actual": team_actual,
        "team_target": team_target,
        "snu_long": snu_long,
        "account_long": account_long,
        "meto_long": meto_long,
    }


def team_target_projection(team_actual: pd.DataFrame, team_target: pd.DataFrame, overall_forecast: pd.DataFrame) -> pd.DataFrame:
    actual_2026 = (
        team_actual.groupby(["본부", "부서"], dropna=False)["sales"]
        .sum()
        .reset_index(name="actual_2026_1_4")
    )
    target_2026 = (
        team_target[team_target["month"] <= "2026-04-01"]
        .groupby(["본부", "부서"], dropna=False)["target"]
        .sum()
        .reset_index(name="target_2026_1_4")
    )
    future_target = (
        team_target[team_target["month"] >= FORECAST_START]
        .groupby(["본부", "부서"], dropna=False)["target"]
        .sum()
        .reset_index(name="target_2026_5_12")
    )
    out = actual_2026.merge(target_2026, how="outer").merge(future_target, how="outer")
    # The blank department row is the headquarters total. Keep team ranking focused on actual teams.
    out = out[out["부서"].notna()].copy()
    out["jan_apr_ar"] = out["actual_2026_1_4"] / out["target_2026_1_4"]
    out["target_adjusted_projection"] = out["target_2026_5_12"] * out["jan_apr_ar"]
    total_ridge = overall_forecast["forecast_sales"].sum()
    out["actual_share_2026_1_4"] = out["actual_2026_1_4"] / out["actual_2026_1_4"].sum()
    out["target_adjusted_share"] = out["target_adjusted_projection"] / out["target_adjusted_projection"].sum()
    out["hybrid_share"] = out["actual_share_2026_1_4"] * 0.75 + out["target_adjusted_share"] * 0.25
    out["ridge_scaled_projection"] = out["hybrid_share"] * total_ridge
    return out.sort_values("ridge_scaled_projection", ascending=False)


def validation_rows(data: dict[str, pd.DataFrame], longs: dict[str, pd.DataFrame]) -> pd.DataFrame:
    rows = []
    checks = [
        ("품목별 실적", data["item_yearly"], longs["item_long"]),
        ("병원별 실적", data["hospital_yearly"], longs["hospital_long"]),
    ]
    for name, src, long_df in checks:
        month_cols = [c for c in src.columns if re.fullmatch(r"\d{1,2}월 실적", str(c))]
        src_sum = sum(pd.to_numeric(src[c].map(clean_number), errors="coerce").sum() for c in month_cols)
        long_sum = long_df["sales"].sum()
        rows.append(
            {
                "check": f"{name} 월별 컬럼 합계 = long 변환 합계",
                "source_sum": src_sum,
                "converted_sum": long_sum,
                "difference": long_sum - src_sum,
                "pass": abs(long_sum - src_sum) < 1e-6,
            }
        )
    for key, long_df in longs.items():
        if "month" not in long_df.columns or long_df.empty:
            continue
        rows.append(
            {
                "check": f"{key} 기간",
                "source_sum": long_df["sales"].sum() if "sales" in long_df.columns else long_df.get("target", pd.Series(dtype=float)).sum(),
                "converted_sum": len(long_df),
                "difference": 0,
                "pass": True,
                "min_month": long_df["month"].min().strftime("%Y-%m"),
                "max_month": long_df["month"].max().strftime("%Y-%m"),
            }
        )
    return pd.DataFrame(rows)


def write_outputs(
    data: dict[str, pd.DataFrame],
    longs: dict[str, pd.DataFrame],
    results: dict[str, pd.DataFrame],
    model: ModelResult,
) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    xlsx_path = OUTPUT_DIR / "ml_forecast_2026_ridge.xlsx"
    report_path = OUTPUT_DIR / "ml_forecast_2026_ridge_report.md"
    code_copy_path = OUTPUT_DIR / "ml_forecast_2026_ridge.py"

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        results["summary"].to_excel(writer, sheet_name="요약", index=False)
        results["overall_forecast"].to_excel(writer, sheet_name="전체예측_2026남은달", index=False)
        results["model_scores"].to_excel(writer, sheet_name="모델검증", index=False)
        results["item_scores"].to_excel(writer, sheet_name="품목별예측요약", index=False)
        results["item_forecast"].to_excel(writer, sheet_name="품목별월별예측", index=False)
        results["hospital_scores"].to_excel(writer, sheet_name="병원별예측요약", index=False)
        results["hospital_forecast"].to_excel(writer, sheet_name="병원별월별예측", index=False)
        results["snu_scores"].to_excel(writer, sheet_name="서울대예측요약", index=False)
        results["snu_forecast"].to_excel(writer, sheet_name="서울대월별예측", index=False)
        results["account_scores"].to_excel(writer, sheet_name="거래처예측요약", index=False)
        results["meto_scores"].to_excel(writer, sheet_name="메토젝트예측요약", index=False)
        results["team_projection"].to_excel(writer, sheet_name="팀별목표보조예측", index=False)
        results["validation"].to_excel(writer, sheet_name="검산", index=False)

    # Add light formatting and a chart with openpyxl after pandas writes the workbook.
    from openpyxl import load_workbook
    from openpyxl.chart import LineChart, Reference
    from openpyxl.styles import Font, PatternFill

    wb = load_workbook(xlsx_path)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
        for col in ws.columns:
            max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 10), 36)
    ws = wb["전체예측_2026남은달"]
    chart = LineChart()
    chart.title = "2026년 남은 달 Ridge 예측"
    chart.y_axis.title = "실적"
    chart.x_axis.title = "월"
    data_ref = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
    cats_ref = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.height = 8
    chart.width = 18
    ws.add_chart(chart, "E2")
    wb.save(xlsx_path)

    summary = results["summary"].iloc[0].to_dict()
    top_items = results["item_scores"].head(10)
    lines = [
        "# 2026년 남은 달 Ridge 예측 리포트",
        "",
        "## 핵심 요약",
        f"- 사용 데이터: 사용자가 제공한 로컬 파일 5개와 data (2).zip 내부 파일 2개만 사용",
        f"- 학습 실적 기간: {summary['actual_min_month']} ~ {summary['actual_max_month']}",
        f"- 예측 기간: {summary['forecast_start']} ~ {summary['forecast_end']}",
        f"- 전체 2026년 1~4월 실적: {summary['actual_2026_1_4']:,.2f}",
        f"- 전체 2026년 5~12월 Ridge 예측: {summary['forecast_2026_5_12']:,.2f}",
        f"- 선택 모델: {summary['best_model']}, 평균 절대 백분율 오차(MAPE): {summary['mape']:.2%}",
        "- 세부 그룹 예측에는 Ridge 외삽이 비정상적으로 0 또는 과대값으로 튀는 것을 막기 위해 최근 평균/전년동월 기준 가드레일을 적용했습니다.",
        "- 팀별 보조 예측은 달성률만 보지 않고 2026년 1~4월 실제 판매량 비중 75%와 목표/달성률 보정 25%를 섞어 배분했습니다.",
        "",
        "## 품목별 예측 상위 10",
        simple_md_table(top_items[["group", "recent_2026_1_4_actual", "forecast_2026_5_12", "model", "mape"]]),
        "",
        "## 검산",
        "- 원자료 월별 실적 합계와 long 변환 후 합계를 비교하는 검산 시트를 포함했습니다.",
        "- 2026년 5~12월은 제공 파일 안에 실제값이 없으므로 예측값입니다.",
    ]
    report_path.write_text("\n".join(lines), encoding="utf-8")
    code_copy_path.write_text(Path(__file__).read_text(encoding="utf-8"), encoding="utf-8")


def main() -> None:
    data = load_sources()
    longs = build_all_long(data)

    main_monthly = aggregate_monthly(longs["hospital_long"])
    overall_forecast, best = forecast_series(main_monthly, FORECAST_MONTHS)

    actual_history = main_monthly[main_monthly["month"] < FORECAST_START].copy()
    featured = add_features(actual_history)
    model_df = featured.dropna(subset=best.feature_cols + ["sales"]).copy()
    fitted = best.model.predict(model_df[best.feature_cols])
    model_scores = pd.DataFrame(
        [
            {
                "model": best.name,
                "MAE": best.mae,
                "MAPE": best.mape,
                "R2": best.r2,
                "train_rows_after_refit": best.train_rows,
                "feature_cols": ", ".join(best.feature_cols),
                "fitted_total_abs_error": float(np.abs(model_df["sales"].to_numpy() - fitted).sum()),
            }
        ]
    )

    item_forecast, item_scores = forecast_groups(
        longs["item_long"],
        ["병원구분", "브랜드", "기준규격"],
        min_months=18,
        top_n=50,
        source_label="품목별 실적",
    )
    hospital_forecast, hospital_scores = forecast_groups(
        longs["hospital_long"],
        ["병원구분", "병원"],
        min_months=18,
        top_n=50,
        source_label="병원별 실적",
    )
    snu_forecast, snu_scores = forecast_groups(
        longs["snu_long"],
        ["구분", "브랜드", "기준규격"],
        min_months=10,
        top_n=30,
        source_label="서울대 품목별 실적",
    )
    account_forecast, account_scores = forecast_groups(
        longs["account_long"],
        ["병원", "팀"],
        min_months=6,
        top_n=30,
        source_label="거래처 실적 순위",
    )
    meto_forecast, meto_scores = forecast_groups(
        longs["meto_long"],
        ["병원", "팀"],
        min_months=6,
        top_n=30,
        source_label="메토젝트 종병 거래처 실적 순위",
    )
    team_projection = team_target_projection(longs["team_actual"], longs["team_target"], overall_forecast)
    validation = validation_rows(data, longs)

    summary = pd.DataFrame(
        [
            {
                "actual_min_month": main_monthly["month"].min().strftime("%Y-%m"),
                "actual_max_month": main_monthly[main_monthly["month"] < FORECAST_START]["month"].max().strftime("%Y-%m"),
                "forecast_start": FORECAST_START.strftime("%Y-%m"),
                "forecast_end": FORECAST_END.strftime("%Y-%m"),
                "actual_2026_1_4": main_monthly[(main_monthly["month"] >= "2026-01-01") & (main_monthly["month"] <= "2026-04-01")]["sales"].sum(),
                "forecast_2026_5_12": overall_forecast["forecast_sales"].sum(),
                "best_model": best.name,
                "mape": best.mape,
                "mae": best.mae,
                "r2": best.r2,
            }
        ]
    )

    overall_forecast_out = overall_forecast.copy()
    overall_forecast_out["month"] = overall_forecast_out["month"].dt.strftime("%Y-%m")
    for df in [item_forecast, hospital_forecast, snu_forecast, account_forecast, meto_forecast]:
        if not df.empty:
            df["month"] = pd.to_datetime(df["month"]).dt.strftime("%Y-%m")

    results = {
        "summary": summary,
        "overall_forecast": overall_forecast_out,
        "model_scores": model_scores,
        "item_forecast": item_forecast,
        "item_scores": item_scores,
        "hospital_forecast": hospital_forecast,
        "hospital_scores": hospital_scores,
        "snu_forecast": snu_forecast,
        "snu_scores": snu_scores,
        "account_forecast": account_forecast,
        "account_scores": account_scores,
        "meto_forecast": meto_forecast,
        "meto_scores": meto_scores,
        "team_projection": team_projection,
        "validation": validation,
    }
    write_outputs(data, longs, results, best)

    summary_json = {
        **summary.iloc[0].to_dict(),
        "workbook": str(OUTPUT_DIR / "ml_forecast_2026_ridge.xlsx"),
        "report": str(OUTPUT_DIR / "ml_forecast_2026_ridge_report.md"),
        "code": str(OUTPUT_DIR / "ml_forecast_2026_ridge.py"),
    }
    (OUTPUT_DIR / "ml_forecast_2026_ridge_summary.json").write_text(
        json.dumps(summary_json, ensure_ascii=False, indent=2, default=str),
        encoding="utf-8",
    )
    print(json.dumps(summary_json, ensure_ascii=False, indent=2, default=str))


if __name__ == "__main__":
    main()
